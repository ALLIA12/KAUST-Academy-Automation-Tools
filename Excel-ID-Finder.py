import openpyxl
import re


def extract_numbers(text):
    return re.findall(r'\d{10}', text)


def read_excel1(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    numbers = set()
    for cell in sheet['E']:
        numbers.update(extract_numbers(str(cell.value)))
    return numbers


def compare_excel_files(file1_path, file2_path):
    numbers_from_file1 = read_excel1(file1_path)

    wb2 = openpyxl.load_workbook(file2_path)
    sheet2 = wb2.active

    missing_names = []

    for row in sheet2.iter_rows(min_row=2, values_only=True):
        number = row[4]  # Column E
        name = row[3]  # Column D

        if str(number) not in numbers_from_file1:
            missing_names.append(name)

    return missing_names


# Usage
file1_path = 'excel 3.xlsx'
file2_path = 'excel 2.xlsx'


missing_names = compare_excel_files(file1_path, file2_path)

print("Names not present in Excel 1:")
for name in missing_names:
    print(name)